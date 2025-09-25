from __future__ import annotations

import os
import numpy as np
import pandas as pd
from datetime import datetime, timedelta, timezone
import openpyxl

OUT_DIR = os.path.join('data', 'GIMtec', 'drivers')
INTERVAL_MIN = 120
STEP_SEC = INTERVAL_MIN * 60


def load_kp_2015_records(path: str):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    year_idx, doy_idx, hr_idx, val_idx = 0, 1, 2, 3
    for r in rows[1:]:
        y, doy, hr, val = r[year_idx], r[doy_idx], r[hr_idx], r[val_idx]
        try:
            y = int(y); doy = int(doy); hr = int(hr)
        except Exception:
            continue
        try:
            v = float(val) / 10.0
        except Exception:
            continue
        dt = datetime(y, 1, 1, tzinfo=timezone.utc) + timedelta(days=doy - 1, hours=hr)
        yield dt, v


def load_dst_2015_records(path: str):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    date_idx, time_idx, dst_idx = 0, 1, 3
    for r in rows[1:]:
        d, t, dst = r[date_idx], r[time_idx], r[dst_idx]
        if d is None or t is None:
            continue
        try:
            dt = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
            dt = dt.replace(hour=getattr(t, 'hour', 0), minute=getattr(t, 'minute', 0))
            v = float(dst)
        except Exception:
            continue
        yield dt, v


def main():
    base = os.path.join('data', 'GIMtec')
    os.makedirs(OUT_DIR, exist_ok=True)

    start = datetime(2009, 1, 1, tzinfo=timezone.utc)
    end = datetime(2022, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
    # build times array
    times = []
    dts = []
    cur = start
    while cur <= end:
        times.append(int(cur.timestamp()))
        dts.append(cur)
        cur = cur + timedelta(seconds=STEP_SEC)
    times = np.array(times, dtype=np.int64)
    T = times.shape[0]

    AE = np.zeros((T,), dtype=np.float32)
    Dst = np.zeros((T,), dtype=np.float32)
    Kp = np.zeros((T,), dtype=np.float32)
    F107 = np.zeros((T,), dtype=np.float32)

    def to_index(dt: datetime) -> int:
        return int(round((dt - start).total_seconds() / STEP_SEC))

    kp_path = os.path.join(base, 'KP_2015.xlsx')
    if os.path.exists(kp_path):
        for dt, v in load_kp_2015_records(kp_path):
            idx = to_index(dt)
            if 0 <= idx < T:
                Kp[idx] = float(v)

    dst_path = os.path.join(base, 'DST_2015.xlsx')
    if os.path.exists(dst_path):
        for dt, v in load_dst_2015_records(dst_path):
            idx = to_index(dt)
            if 0 <= idx < T:
                Dst[idx] = float(v)

    # Compute cosSZA over 71x73 grid for full timeline (pure Python datetime → numpy)
    day_of_year = np.array([dt.timetuple().tm_yday for dt in dts], dtype=np.float32)
    gamma = 2.0 * np.pi * (day_of_year - 1) / 365.0
    decl = (
        0.006918 - 0.399912 * np.cos(gamma) + 0.070257 * np.sin(gamma)
        - 0.006758 * np.cos(2 * gamma) + 0.000907 * np.sin(2 * gamma)
        - 0.002697 * np.cos(3 * gamma) + 0.00148 * np.sin(3 * gamma)
    )
    eqtime = 229.18 * (
        0.000075 + 0.001868 * np.cos(gamma) - 0.032077 * np.sin(gamma)
        - 0.014615 * np.cos(2 * gamma) - 0.040849 * np.sin(2 * gamma)
    )
    lon = np.linspace(0.0, 360.0 - 5.0, 73, dtype=np.float32)
    lat = np.linspace(-87.5, 87.5, 71, dtype=np.float32)
    Lon, Lat = np.meshgrid(lon, lat)
    hour = np.array([dt.hour for dt in dts], dtype=np.float32)
    minute = np.array([dt.minute for dt in dts], dtype=np.float32)
    Ht = (np.expand_dims(hour * 60 + minute + eqtime, 1) + 4.0 * Lon.reshape(1, -1))
    Ht = np.deg2rad(Ht / 4.0 - 180.0)
    latr = np.deg2rad(Lat.reshape(1, -1))
    cosSZA = np.sin(latr) * np.sin(decl).reshape(-1, 1) + np.cos(latr) * np.cos(decl).reshape(-1, 1) * np.cos(Ht)
    cosSZA = np.clip(cosSZA, 0.0, 1.0).astype(np.float32)  # [T, 5183]

    out_path = os.path.join(OUT_DIR, 'drivers_2009_2022_2h.npz')
    np.savez_compressed(out_path, time=times, AE=AE, Dst=Dst, Kp=Kp, F107=F107, cosSZA=cosSZA,
                        meta=dict(interval_min=INTERVAL_MIN, grid='71x73'))
    print('Saved real drivers (partial 2015):', out_path)


if __name__ == '__main__':
    main()
